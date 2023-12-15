import time
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
import bs4
import requests
import shutil
from private import CSV_PATH, OUTPUT_PATH, BACKUP_PATH, INVESTMENT_TRUST_NAMES


class InvestmentTrustData:

    def __init__(self):
        # 基準日
        self.base_date = ''
        # 基準価格
        self.today_price = ''
        # 前日差額
        self.before_price = ''
        # 前日比
        self.before_ratio = ''
        # 純資産
        self.total_assets = 0


URLS = ['https://site0.sbisec.co.jp/marble/fund/detail/achievement.do?Param6=10331418A',
        'https://site0.sbisec.co.jp/marble/fund/detail/achievement.do?Param6=103311187',
        'https://site0.sbisec.co.jp/marble/fund/detail/achievement.do?Param6=20331A211']
DOWNLOADS = ['https://www.am.mufg.jp/fund_file/setteirai/253425.csv',
             'https://www.am.mufg.jp/fund_file/setteirai/253266.csv',
             'https://www.am.mufg.jp/fund_file/setteirai/254062.csv']

# 投資信託CSVを取得し特定フォルダに保存
for i in range(0, len(DOWNLOADS)):
    res = requests.get(DOWNLOADS[i])
    res.raise_for_status()
    f = open(f'{CSV_PATH}/{INVESTMENT_TRUST_NAMES[i]}.csv', 'wb')
    for chunk in res.iter_content(100000):
        f.write(chunk)
    f.close()
    time.sleep(1)

# スクレイピングしたデータの保存用
datas = []

# 当日の値をスクレイピングで取得する
for url in URLS:
    data = InvestmentTrustData()
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.content, 'html.parser')
    # 基準日(YY/MM/DD)
    date = soup.select('div.floatL.fGray01.md-l-utl-mt9')[0].text.strip()[1:9]
    data.base_date = '20' + date
    # 基準価格
    price = soup.select('div.floatL.bold')[0].text.strip()[0:6]
    data.today_price = price
    # 前日比
    before_info = soup.select('div.floatL.md-l-utl-mt9 > span')
    change_price = before_info[0].text.strip()[:-1]
    change_ratio = before_info[1].text.strip().replace('％', '%')
    data.before_price = change_price
    data.before_ratio = change_ratio
    # 総資産
    total_assets = soup.select('table.md-l-table-01.has_tooltip.lower td.alR')[0].text.strip()
    # カンマの除去して単位を百万から億に変更
    data.total_assets = int(total_assets[:-3].replace(',', '')) / 100
    datas.append(data)
    time.sleep(1)

# 取得した情報をエクセルに書き込み保存
file_name = f'{datas[0].base_date[:4]}_投資信託.xlsx'
wb = openpyxl.load_workbook(f'{OUTPUT_PATH}/{file_name}')
i = 0
right = Alignment(horizontal='right')
is_modify = False
for name in INVESTMENT_TRUST_NAMES:
    sheet = wb[name]
    insert_row = sheet.max_row + 1
    data = datas[i]
    before_date = sheet.cell(row=sheet.max_row, column=1).value
    # 一番下の行の日付とスクレイピングしたデータの日付が同じ場合、エクセルに挿入しない
    if data.base_date == before_date:
        break
    is_modify = True
    # 基準日
    sheet.cell(row=insert_row, column=1).value = data.base_date
    # 基準価格
    sheet.cell(row=insert_row, column=2).value = data.today_price
    sheet.cell(row=insert_row, column=2).alignment = right
    # 前日比
    before = f'{data.before_price}({data.before_ratio})'
    sheet.cell(row=insert_row, column=3).value = before
    sheet.cell(row=insert_row, column=3).alignment = right
    if before[0] == '+':
        sheet.cell(row=insert_row, column=3).font = Font(color='FF0000')
    elif before[0] == '-':
        sheet.cell(row=insert_row, column=3).font = Font(color='0000FF')
    # 純資産
    sheet.cell(row=insert_row, column=4).value = data.total_assets
    i += 1

if BACKUP_PATH.exists():
    shutil.rmtree(BACKUP_PATH)
# エクセルに変更があった場合、保存、バックアップにコピーする
if is_modify:
    shutil.copytree(OUTPUT_PATH, BACKUP_PATH)
    wb.save(f'Output/{file_name}')
