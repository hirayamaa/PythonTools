import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
import csv
import datetime
from private import CSV_PATH, INVESTMENT_TRUST_NAMES


def get_day_before_ratio(today_price, before_price):
    if before_price == 0:
        return ''
    change_price = today_price - before_price
    plus_str = ''
    if change_price > 0:
        plus_str = '+'
    up_rate = round((today_price / before_price - 1) * 100, 2)
    return f'{plus_str}{change_price}({plus_str}{up_rate:3.2f}%)'


# 設定来データをエクセルに書き込む(2018年~現在年)
year_range = range(2018, int(datetime.date.today().year) + 1)
for year in year_range:
    # 年ごとのエクセルファイルを作成する(ファイル名:YYYY_投資信託.xlsx)
    wb = openpyxl.Workbook()
    # 投資信託銘柄用のシートをそれぞれ作成する
    for name in INVESTMENT_TRUST_NAMES:
        wb.create_sheet(title=name)
        # ヘッダーを書き込む
        sheet = wb[name]
        sheet.cell(row=1, column=1).value = '基準日'
        sheet.cell(row=1, column=2).value = '基準価格(円)'
        sheet.cell(row=1, column=3).value = '前日比'
        sheet.cell(row=1, column=4).value = '純資産総額(億円)'
        # 各列の幅調整
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 16
        # 年初来データのCSVを読み込み、各シートに書き込む
        csv_file = open(f'{CSV_PATH}/{name}.csv', encoding='ShiftJIS')
        csv_reader = csv.reader(csv_file)

        insert_row_num = 2
        yesterday_price = 0
        right = Alignment(horizontal='right')
        for row in csv_reader:
            # 基準日が書き込み対象の年でない場合、処理なし
            if row[0][:4] != str(year):
                continue
            # 基準日
            sheet.cell(row=insert_row_num, column=1).value = row[0]
            # 基準価格
            sheet.cell(row=insert_row_num, column=2).value = format(int(row[1]), ',')
            sheet.cell(row=insert_row_num, column=2).alignment = right
            # 前日比
            sheet.cell(row=insert_row_num, column=3).value = get_day_before_ratio(int(row[1]), int(yesterday_price))
            sheet.cell(row=insert_row_num, column=3).alignment = right
            # 前日比がプラスなら赤文字、マイナスなら青文字にする
            if int(row[1]) - int(yesterday_price) > 0:
                sheet.cell(row=insert_row_num, column=3).font = Font(color='FF0000')
            elif int(row[1]) - int(yesterday_price) < 0:
                sheet.cell(row=insert_row_num, column=3).font = Font(color='0000FF')
            # 純資産
            sheet.cell(row=insert_row_num, column=4).value = row[4]
            sheet.cell(row=insert_row_num, column=4).alignment = right
            insert_row_num += 1
            yesterday_price = row[1]
    del wb['Sheet']
    wb.save(f'Output/{year}_投資信託.xlsx')
